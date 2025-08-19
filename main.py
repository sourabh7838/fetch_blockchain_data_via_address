import requests
import pandas as pd
import numpy as np
import time
import openpyxl
import os

################### API configuration for Blockchain.info API ####################
API_KEY = "" ## Please add your API Key ex:- 'd20c0000-0000-0000-0000-d0ad22ac0000'
BLOCKCHAIN_CONFIG = {
    'base_delay': 1,           # Base delay between requests (sec.)
    'inter_address_delay': 3,  # Delay b/w different addresses
    'max_retries': 10,         # Max retry attempts
    'timeout': 30              # Request timeout
}

#################### one Bitcoin addresses for test the functionality ###################
bitcoin_addresses = [
    '1KFHE7w8BhaENAswwryaoccDb6qcT6DbYY',
]

#################### Setup and API Interaction ###################

def validate_api_connectivity():
    """
    Validates connectivity to Blockchain.info API.
    """
    try:
        #################### Test with a simple API endpoint ###################
        test_url = "https://blockchain.info/rawaddr/1A1zP1eP5QGefi2DMPTfTL5SLmv7DivfNa"
        params = {'limit': 1, 'api_code': API_KEY}
        response = requests.get(test_url, params=params, timeout=10)
        return response.status_code == 200
    except:
        return False

def get_user_choice():
    """
    Get user's choice on how to proceed with analysis.
    """
    print("\n Choose your analysis option:")
    print("1. Test and Analyze default addresses")
    print("2. Use addresses from addresses.txt file")
    print("3. Process just one address for testing")
    print("4. Exit")
    
    while True:
        try:
            choice = input("\nEnter your choice (1-4): ").strip()
            if choice in ['1', '2', '3', '4']:
                return int(choice)
            print("Please enter 1, 2, 3, or 4")
        except KeyboardInterrupt:
            print("\n\nExiting...")
            return 4

def fetch_address_data(address):
    """
    Fetches transaction data using Blockchain.info API.
    
    Args:
        address (str): The Bitcoin address to query.
        
    Returns:
        tuple: (transactions, address_data) or (None, None) if error occurs.
    """
    print(f" Fetching data for {address} using Blockchain.info API...")
    
    #################### Get address data with transactions
    address_url = f"https://blockchain.info/rawaddr/{address}"
    params = {
        'limit': 1000,  # Get up to 1000 transactions - u can change it
        'api_code': API_KEY
    }
    
    retries = 0
    while retries < BLOCKCHAIN_CONFIG['max_retries']:
        try:
            if retries > 0:
                delay = BLOCKCHAIN_CONFIG['base_delay'] * (retries + 1)
                print(f" Waiting {delay} seconds before retry...")
                time.sleep(delay)
            else:
                time.sleep(BLOCKCHAIN_CONFIG['base_delay'])
            
            #################### Get address information and transactions
            response = requests.get(address_url, params=params, timeout=BLOCKCHAIN_CONFIG['timeout'])
            
            if response.status_code == 200:
                address_data = response.json()
                transactions = address_data.get('txs', [])
                
                print(f" Retrieved {len(transactions)} transactions for {address}")
                
                balance = address_data.get('final_balance', 0) / 1e8
                total_received = address_data.get('total_received', 0) / 1e8
                total_sent = address_data.get('total_sent', 0) / 1e8
                tx_count = address_data.get('n_tx', 0)
                
                print(f" Address info: Balance: {balance:.8f} BTC, Total RX: {total_received:.8f} BTC, Total TX: {tx_count}")
                
                return transactions, address_data
                
            elif response.status_code == 429:
                print(f"  Rate limit hit. Retry {retries + 1}/{BLOCKCHAIN_CONFIG['max_retries']}")
                retries += 1
                continue
                
            else:
                print(f" HTTP {response.status_code}: {response.text[:200]}")
                retries += 1
                continue
                
        except requests.exceptions.RequestException as e:
            print(f" Network error: {e}")
            retries += 1
            if retries < BLOCKCHAIN_CONFIG['max_retries']:
                time.sleep(3 * retries)
    
    print(f" Failed to fetch data for {address} after {BLOCKCHAIN_CONFIG['max_retries']} retries")
    return None, None

def process_transactions(transactions, target_address):
    """
    Separates transactions into 'input' and 'output' based on the target address.
    This version works with Blockchain.info API format.
    
    Args:
        transactions (list): A list of transaction dictionaries from Blockchain.info API.
        target_address (str): The Bitcoin address being analyzed.
        
    Returns:
        tuple: A tuple containing two lists: (in_txs, out_txs).
    """
    in_txs = []  ####### Transactions where the target_address is a sender (in input)
    out_txs = [] ####### Transactions where the target_address is a receiver (in output)

    for tx in transactions:
        # Check inputs (Blockchain.info API format)
        is_in_tx = False
        for input_tx in tx.get('inputs', []):
            if input_tx.get('prev_out', {}).get('addr') == target_address:
                is_in_tx = True
                break
        
        # Check outputs (Blockchain.info API format)
        is_out_tx = False
        for output in tx.get('out', []):
            if output.get('addr') == target_address:
                is_out_tx = True
                break
        
        if is_in_tx:
            in_txs.append(tx)
        if is_out_tx:
            out_txs.append(tx)

    return in_txs, out_txs

############################### Calculation Functions #################################

def calculate_in_params(in_txs, target_address):
    """
    Calculates parameters for input transactions using Blockchain.info API data structure.
    
    Args:
        in_txs (list): A list of input transaction dictionaries.
        target_address (str): The Bitcoin address being analyzed.
        
    Returns:
        dict: A dictionary of calculated input parameters.
    """
    if not in_txs:
        return {
            '1. No. of in. transactions': 0,
            '2. Total recipient addresses (excluding self as change)': 0,
            '3. Number of unique recipient addresses': 0,
            '4. Average number of recipients per transaction': 0,
            '5. Max number of recipients in a transaction': 0,
            '6. Min number of recipients in a transaction': 0,
            '7. Standard deviation of recipients counts': 0,
            '8. Total sender addresses in in. transactions': 0,
            '9. Number of unique senders addresses': 0,
            '10. Average number of senders per transaction': 0,
            '11. Max number of senders in a transaction': 0,
            '12. Min number of senders in a transaction': 0,
            '13. Standard deviation of sender counts': 0,
            '14. Total coins transferred (excluding change)': 0,
            '15. Average coins transferred per transaction': 0,
            '16. Min coins transferred in one transaction': 0,
            '17. Max coins transferred in one transaction': 0,
            '18. Standard deviation of coins transferred in all transaction': 0,
            '19. Avg. coins transferred per receiver': 0,
            '20. Avg. coins transferred per unique receiver': 0
        }

    recipients_counts = []
    sender_counts = []
    transferred_coins = []
    all_recipients = []
    all_senders = set()
    
    for tx in in_txs:

        current_senders = []
        for input_tx in tx.get('inputs', []):
            addr = input_tx.get('prev_out', {}).get('addr')
            if addr:
                current_senders.append(addr)
        all_senders.update(current_senders)
        sender_counts.append(len(current_senders))
        

        current_recipients = []
        for output in tx.get('out', []):
            addr = output.get('addr')
            if addr and addr != target_address:
                current_recipients.append(addr)
        recipients_counts.append(len(current_recipients))
        all_recipients.extend(current_recipients)
        
        sent_amount = 0
        for output in tx.get('out', []):
            if output.get('addr') != target_address:
                sent_amount += output.get('value', 0)
        transferred_coins.append(sent_amount)

    transferred_coins_btc = [c / 1e8 for c in transferred_coins]
    
    avg_coins_per_receiver = 0
    if all_recipients:
        total_coins = sum(transferred_coins)
        avg_coins_per_receiver = (total_coins / len(all_recipients)) / 1e8


    avg_coins_per_unique_receiver = 0
    unique_recipients = set(all_recipients)
    if unique_recipients:
        total_coins = sum(transferred_coins)
        avg_coins_per_unique_receiver = (total_coins / len(unique_recipients)) / 1e8
        
    return {

        '1. No. of in. transactions': len(in_txs),
        '2. Total recipient addresses (excluding self as change)': len(all_recipients),
        '3. Number of unique recipient addresses': len(unique_recipients),
        '4. Average number of recipients per transaction': np.mean(recipients_counts) if recipients_counts else 0,
        '5. Max number of recipients in a transaction': np.max(recipients_counts) if recipients_counts else 0,
        '6. Min number of recipients in a transaction': np.min(recipients_counts) if recipients_counts else 0,
        '7. Standard deviation of recipients counts': np.std(recipients_counts) if len(recipients_counts) > 1 else 0,
        '8. Total sender addresses in in. transactions': len(all_senders),
        '9. Number of unique senders addresses': len(all_senders),
        '10. Average number of senders per transaction': np.mean(sender_counts) if sender_counts else 0,
        '11. Max number of senders in a transaction': np.max(sender_counts) if sender_counts else 0,
        '12. Min number of senders in a transaction': np.min(sender_counts) if sender_counts else 0,
        '13. Standard deviation of sender counts': np.std(sender_counts) if len(sender_counts) > 1 else 0,
        '14. Total coins transferred (excluding change)': sum(transferred_coins_btc),
        '15. Average coins transferred per transaction': np.mean(transferred_coins_btc) if transferred_coins_btc else 0,
        '16. Min coins transferred in one transaction': np.min(transferred_coins_btc) if transferred_coins_btc else 0,
        '17. Max coins transferred in one transaction': np.max(transferred_coins_btc) if transferred_coins_btc else 0,
        '18. Standard deviation of coins transferred in all transaction': np.std(transferred_coins_btc) if len(transferred_coins_btc) > 1 else 0,
        '19. Avg. coins transferred per receiver': avg_coins_per_receiver,
        '20. Avg. coins transferred per unique receiver': avg_coins_per_unique_receiver
    }

def calculate_out_params(out_txs, target_address):
    """
    Calculates parameters for output transactions using Blockchain.info API data structure.
    
    Args:
        out_txs (list): A list of output transaction dictionaries.
        target_address (str): The Bitcoin address being analyzed.
        
    Returns:
        dict: A dictionary of calculated output parameters.
    """
    if not out_txs:
        return {
            '21. Number of out. transactions': 0,
            '22. Total senders addresses (excluding self as change)': 0,
            '23. Number of unique sender addresses': 0,
            '24. Average number of senders per transaction': 0,
            '25. Max number of senders in a transaction': 0,
            '26. Min number of senders in a transaction': 0,
            '27. Variation (std. dev.) in senders count': 0,
            '28. Total receivers addresses in out. transactions': 0,
            '29. Number of unique receivers addresses': 0,
            '30. Average number of receivers per transaction': 0,
            '31. Max number of receivers in a transaction': 0,
            '32. Min number of receivers in a transaction': 0,
            '33. Variation in receivers count': 0,
            '34. Total coins received (excluding change)': 0,
            '35. Average coins received per transaction': 0,
            '36. Min coins received in one transaction': 0,
            '37. Max coins received in one transaction': 0,
            '38. Variation (i.e., S.D) in coins received in all transaction': 0,
            '39. Avg. coins per sender': 0
        }

    senders_counts = []
    receivers_counts = []
    received_coins = []
    all_senders = set()
    all_receivers = []
    
    for tx in out_txs:

        current_senders = []
        for input_tx in tx.get('inputs', []):
            addr = input_tx.get('prev_out', {}).get('addr')
            if addr and addr != target_address:  # Exclude self as change
                current_senders.append(addr)
        senders_counts.append(len(current_senders))
        all_senders.update(current_senders)
        
        current_receivers = []
        for output in tx.get('out', []):
            addr = output.get('addr')
            if addr:
                current_receivers.append(addr)
        receivers_counts.append(len(current_receivers))
        all_receivers.extend(current_receivers)
        
        amount_received = 0
        for output in tx.get('out', []):
            if output.get('addr') == target_address:
                amount_received += output.get('value', 0)
        received_coins.append(amount_received)

    #################### Convert to Bitcoin (Satoshi to BTC) ###################
    received_coins_btc = [c / 1e8 for c in received_coins]
    
    #################### Calculate average coins per sender ###################
    avg_coins_per_sender = 0
    if all_senders:
        total_coins = sum(received_coins)
        avg_coins_per_sender = (total_coins / len(all_senders)) / 1e8
        
    return {
        #################### Output transaction parameters (exactly as per assignment requirements)  ###################
        '21. Number of out. transactions': len(out_txs),
        '22. Total senders addresses (excluding self as change)': len(all_senders),
        '23. Number of unique sender addresses': len(all_senders),
        '24. Average number of senders per transaction': np.mean(senders_counts) if senders_counts else 0,
        '25. Max number of senders in a transaction': np.max(senders_counts) if senders_counts else 0,
        '26. Min number of senders in a transaction': np.min(senders_counts) if senders_counts else 0,
        '27. Variation (std. dev.) in senders count': np.std(senders_counts) if len(senders_counts) > 1 else 0,
        '28. Total receivers addresses in out. transactions': len(all_receivers),
        '29. Number of unique receivers addresses': len(set(all_receivers)),
        '30. Average number of receivers per transaction': np.mean(receivers_counts) if receivers_counts else 0,
        '31. Max number of receivers in a transaction': np.max(receivers_counts) if receivers_counts else 0,
        '32. Min number of receivers in a transaction': np.min(receivers_counts) if receivers_counts else 0,
        '33. Variation in receivers count': np.std(receivers_counts) if len(receivers_counts) > 1 else 0,
        '34. Total coins received (excluding change)': sum(received_coins_btc),
        '35. Average coins received per transaction': np.mean(received_coins_btc) if received_coins_btc else 0,
        '36. Min coins received in one transaction': np.min(received_coins_btc) if received_coins_btc else 0,
        '37. Max coins received in one transaction': np.max(received_coins_btc) if received_coins_btc else 0,
        '38. Variation (i.e., S.D) in coins received in all transaction': np.std(received_coins_btc) if len(received_coins_btc) > 1 else 0,
        '39. Avg. coins per sender': avg_coins_per_sender
    }

###################################### Main Execution ######################################

def analyze_addresses(addresses):
    """
    Main function to orchestrate the entire analysis.
    
    Args:
        addresses (list): A list of Bitcoin addresses to analyze.
        
    Returns:
        pandas.DataFrame: A DataFrame with the analysis results for all addresses.
    """
    results_list = []
    total_addresses = len(addresses)
    
    for i, address in enumerate(addresses, 1):
        print(f"\n=== Starting analysis for address {i}/{total_addresses}: {address} ===")
        
        if i > 1:
            inter_address_delay = BLOCKCHAIN_CONFIG['inter_address_delay']
            print(f"Waiting {inter_address_delay} seconds before processing next address...")
            time.sleep(inter_address_delay)
        
        try:
            txs, address_data = fetch_address_data(address)
            
            if txs is None:
                print(f" Skipping analysis for {address} due to data fetching error.")
                continue
            
            print(f" Successfully fetched {len(txs)} transactions for {address}")
            
            in_txs, out_txs = process_transactions(txs, address)
            print(f"    Processing: {len(in_txs)} input transactions, {len(out_txs)} output transactions")
            
            in_params = calculate_in_params(in_txs, address)
            out_params = calculate_out_params(out_txs, address)
            
            balance = address_data.get('final_balance', 0) / 1e8
            total_received = address_data.get('total_received', 0) / 1e8
            total_sent = address_data.get('total_sent', 0) / 1e8
            
            combined_params = {
                'Bitcoin Address': address,
                'Current Balance (BTC)': balance,
                'Total Received (BTC)': total_received,
                'Total Sent (BTC)': total_sent,
                **in_params,
                **out_params
            }
            results_list.append(combined_params)
            print(f" Completed analysis for {address}")
            
        except Exception as e:
            print(f" Unexpected error analyzing {address}: {e}")
            continue
        
    print(f"\n Analysis completed for {len(results_list)}/{total_addresses} addresses")
    return pd.DataFrame(results_list)

def save_to_excel(dataframe, filename='bitcoin_analysis.xlsx'):
    """
    Saves the final DataFrame to a professionally formatted Excel file with multiple sheets.
    
    Args:
        dataframe (pandas.DataFrame): The DataFrame to save.
        filename (str): The name of the output Excel file.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.utils import get_column_letter
        
        #################### Create a new workbook ###################
        wb = Workbook()
        
        #################### Remove default sheet and create named sheets ###################
        wb.remove(wb.active)
        
        #################### Create Summary sheet ###################
        summary_sheet = wb.create_sheet("Summary")
        
        #################### Create detailed analysis sheet ###################
        detail_sheet = wb.create_sheet("Detailed Analysis")
        
        #################### Create parameter definitions sheet ###################
        definitions_sheet = wb.create_sheet("Parameter Guide")
        
        #################### SUMMARY SHEET ###################
        # Main title - no merged cells to avoid issues
        title_cell = summary_sheet['A1']
        title_cell.value = "Bitcoin Address Analysis Summary Report"
        title_cell.font = Font(bold=True, size=18, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        summary_sheet.row_dimensions[1].height = 35
        
        #################### Add analysis timestamp ###################
        from datetime import datetime
        timestamp_cell = summary_sheet['A2']
        timestamp_cell.value = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        timestamp_cell.font = Font(italic=True, size=11, color='666666')
        timestamp_cell.alignment = Alignment(horizontal='left')
        summary_sheet.row_dimensions[2].height = 20
        
        #################### Summary statistics section ###################
        row = 4
        stats_cell = summary_sheet[f'A{row}']
        stats_cell.value = "OVERALL STATISTICS"
        stats_cell.font = Font(bold=True, size=14, color='FFFFFF')
        stats_cell.fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        stats_cell.alignment = Alignment(horizontal='left', vertical='center')
        summary_sheet.row_dimensions[row].height = 25
        row += 2
        
        summary_data = [
            ['Total Addresses Analyzed', len(dataframe)],
            ['Total Current Balance (BTC)', f"{dataframe['Current Balance (BTC)'].sum():.8f}"],
            ['Total Received (BTC)', f"{dataframe['Total Received (BTC)'].sum():.8f}"],
            ['Total Sent (BTC)', f"{dataframe['Total Sent (BTC)'].sum():.8f}"],
            ['Total Input Transactions', int(dataframe['1. No. of in. transactions'].sum())],
            ['Total Output Transactions', int(dataframe['21. Number of out. transactions'].sum())],
            ['Total Transactions Analyzed', int(dataframe['1. No. of in. transactions'].sum() + dataframe['21. Number of out. transactions'].sum())],
        ]
        
        for item in summary_data:
            #################### Label column ###################
            label_cell = summary_sheet[f'A{row}']
            label_cell.value = item[0]
            label_cell.font = Font(bold=True, size=11, color='1F4E79')
            label_cell.fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')
            label_cell.alignment = Alignment(horizontal='left', vertical='center')
            label_cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            #################### Value column ###################
            value_cell = summary_sheet[f'B{row}']
            value_cell.value = item[1]
            if isinstance(item[1], str) and 'BTC' in item[1]:
                value_cell.font = Font(bold=True, color='C55A11', size=11)  # Orange for BTC amounts
            else:
                value_cell.font = Font(bold=True, color='2E75B6', size=11)  # Blue for counts
            value_cell.alignment = Alignment(horizontal='right', vertical='center')
            value_cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            row += 1
        
        #################### Address overview table ###################
        row += 2
        overview_cell = summary_sheet[f'A{row}']
        overview_cell.value = "🏦 ADDRESS OVERVIEW"
        overview_cell.font = Font(bold=True, size=14, color='FFFFFF')
        overview_cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        overview_cell.alignment = Alignment(horizontal='left', vertical='center')
        summary_sheet.row_dimensions[row].height = 25
        row += 2
        
        #################### Headers for address overview ###################
        headers = ['Bitcoin Address', 'Balance (BTC)', 'Total Received (BTC)', 'Total Sent (BTC)', 'Input TXs', 'Output TXs']
        for col, header in enumerate(headers, 1):
            cell = summary_sheet.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        
        row += 1
        for row_idx, (_, address_data) in enumerate(dataframe.iterrows()):
            row_color = 'F8F9FA' if row_idx % 2 == 0 else 'FFFFFF'
            
            addr = address_data['Bitcoin Address']
            display_addr = f"{addr[:12]}...{addr[-8:]}" if len(addr) > 25 else addr
            addr_cell = summary_sheet.cell(row=row, column=1, value=display_addr)
            addr_cell.font = Font(bold=True, color='1F4E79', size=10)
            addr_cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')
            
            balance_value = address_data['Current Balance (BTC)']
            balance_cell = summary_sheet.cell(row=row, column=2, value=f"{balance_value:.8f}")
            if balance_value > 1:
                balance_cell.font = Font(color='C55A11', bold=True, size=10)
            elif balance_value > 0:
                balance_cell.font = Font(color='70AD47', bold=True, size=10)
            else:
                balance_cell.font = Font(color='7C7C7C', size=10)
            balance_cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')
            
            received_cell = summary_sheet.cell(row=row, column=3, value=f"{address_data['Total Received (BTC)']:.8f}")
            received_cell.font = Font(color='70AD47', bold=True, size=10)
            received_cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')
            
            sent_cell = summary_sheet.cell(row=row, column=4, value=f"{address_data['Total Sent (BTC)']:.8f}")
            sent_cell.font = Font(color='C55A11', bold=True, size=10)
            sent_cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')
            
            in_tx_cell = summary_sheet.cell(row=row, column=5, value=int(address_data['1. No. of in. transactions']))
            in_tx_cell.font = Font(color='2E75B6', bold=True, size=10)
            in_tx_cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')
            
            out_tx_cell = summary_sheet.cell(row=row, column=6, value=int(address_data['21. Number of out. transactions']))
            out_tx_cell.font = Font(color='2E75B6', bold=True, size=10)
            out_tx_cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')
            
            #################### Add borders and alignment to all cells ###################
            for col in range(1, 7):
                cell = summary_sheet.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin', color='D0D0D0'), 
                    right=Side(style='thin', color='D0D0D0'),
                    top=Side(style='thin', color='D0D0D0'), 
                    bottom=Side(style='thin', color='D0D0D0')
                )
                if col == 1:  # Address column
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            row += 1
        
        #################### DETAILED ANALYSIS SHEET ###################
        detail_title = detail_sheet['A1']
        detail_title.value = "Detailed Bitcoin Address Analysis - All 39 Parameters"
        detail_title.font = Font(bold=True, size=16, color='FFFFFF')
        detail_title.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        detail_title.alignment = Alignment(horizontal='left', vertical='center')
        detail_sheet.row_dimensions[1].height = 30
        
        #################### Add explanatory note ###################
        detail_note = detail_sheet['A2']
        detail_note.value = "This sheet contains all calculated parameters for each Bitcoin address as per assignment requirements"
        detail_note.font = Font(italic=True, size=11, color='666666')
        detail_note.alignment = Alignment(horizontal='left')
        detail_sheet.row_dimensions[2].height = 20
        
        #################### Add all data to detailed sheet starting from row 4 ###################
        row_offset = 4
        
        #################### Add headers with better formatting ###################
        headers = list(dataframe.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = detail_sheet.cell(row=row_offset, column=col_idx, value=header)
            
            if 'Bitcoin Address' in header or 'Balance' in header or 'Received' in header or 'Sent' in header:
                cell.fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
            elif header.startswith(('1.', '2.', '3.', '4.', '5.', '6.', '7.', '8.', '9.', '10.')):
                cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')  # Input params - Green
            elif header.startswith(('11.', '12.', '13.', '14.', '15.', '16.', '17.', '18.', '19.', '20.')):
                cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')  # Input params - Green
            else:
                cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')  # Output params - Orange
            
            cell.font = Font(bold=True, color='FFFFFF', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        
        for row_idx, (_, row_data) in enumerate(dataframe.iterrows(), row_offset + 1):
            row_color = 'F8F9FA' if (row_idx - row_offset) % 2 == 0 else 'FFFFFF'
            
            for col_idx, value in enumerate(row_data, 1):
                cell = detail_sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin', color='D0D0D0'), 
                    right=Side(style='thin', color='D0D0D0'),
                    top=Side(style='thin', color='D0D0D0'), 
                    bottom=Side(style='thin', color='D0D0D0')
                )
                cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')
                
                if isinstance(value, (int, float)) and col_idx > 1:
                    if 'BTC' in headers[col_idx-1] or any(x in headers[col_idx-1] for x in ['coins', 'transferred', 'received']):
                        cell.number_format = '0.00000000'
                        if value > 0:
                            cell.font = Font(color='C55A11', bold=True, size=9)
                        else:
                            cell.font = Font(color='7C7C7C', size=9)
                    elif isinstance(value, float):
                        cell.number_format = '0.00'
                        cell.font = Font(color='2E75B6', size=9)
                    else:
                        cell.font = Font(color='1F4E79', size=9)
                elif col_idx == 1:
                    cell.font = Font(bold=True, color='1F4E79', size=9)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        
        #################### PARAMETER DEFINITIONS SHEET ###################
        def_title = definitions_sheet['A1']
        def_title.value = "Parameter Definitions & Assignment Requirements"
        def_title.font = Font(bold=True, size=16, color='FFFFFF')
        def_title.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        def_title.alignment = Alignment(horizontal='left', vertical='center')
        definitions_sheet.row_dimensions[1].height = 30
        
        #################### Parameter definitions with assignment requirements ###################
        definitions = [
            ["Parameter Number", "Parameter Name", "Description", "Category"],
            ["", "", "", ""],
            ["INPUT TRANSACTION PARAMETERS (1-20)", "", "", ""],
            ["1", "No. of in. transactions", "Total number of transactions where this address sends coins to others", "Count"],
            ["2", "Total recipient addresses (excluding self as change)", "Total count of addresses that received coins (change to self excluded)", "Count"],
            ["3", "Number of unique recipient addresses", "Count of unique addresses that received coins from this address", "Count"],
            ["4", "Average number of recipients per transaction", "Average number of recipients per input transaction", "Average"],
            ["5", "Max number of recipients in a transaction", "Maximum recipients in any single input transaction", "Maximum"],
            ["6", "Min number of recipients in a transaction", "Minimum recipients in any single input transaction", "Minimum"],
            ["7", "Standard deviation of recipients counts", "Statistical variation in number of recipients per transaction", "Statistics"],
            ["8", "Total sender addresses in in. transactions", "Total count of all sender addresses in input transactions", "Count"],
            ["9", "Number of unique senders addresses", "Count of unique sender addresses in input transactions", "Count"],
            ["10", "Average number of senders per transaction", "Average number of senders per input transaction", "Average"],
            ["11", "Max number of senders in a transaction", "Maximum senders in any single input transaction", "Maximum"],
            ["12", "Min number of senders in a transaction", "Minimum senders in any single input transaction", "Minimum"],
            ["13", "Standard deviation of sender counts", "Statistical variation in number of senders per transaction", "Statistics"],
            ["14", "Total coins transferred (excluding change)", "Total BTC amount sent to others (change to self excluded)", "BTC Amount"],
            ["15", "Average coins transferred per transaction", "Average BTC amount sent per input transaction", "BTC Amount"],
            ["16", "Min coins transferred in one transaction", "Minimum BTC amount sent in any single transaction", "BTC Amount"],
            ["17", "Max coins transferred in one transaction", "Maximum BTC amount sent in any single transaction", "BTC Amount"],
            ["18", "Standard deviation of coins transferred in all transaction", "Statistical variation in BTC amounts sent per transaction", "Statistics"],
            ["19", "Avg. coins transferred per receiver", "Average BTC amount per recipient address", "BTC Amount"],
            ["20", "Avg. coins transferred per unique receiver", "Average BTC amount per unique recipient address", "BTC Amount"],
            ["", "", "", ""],
            ["OUTPUT TRANSACTION PARAMETERS (21-39)", "", "", ""],
            ["21", "Number of out. transactions", "Total number of transactions where this address receives coins", "Count"],
            ["22", "Total senders addresses (excluding self as change)", "Total count of addresses that sent coins (self excluded)", "Count"],
            ["23", "Number of unique sender addresses", "Count of unique addresses that sent coins to this address", "Count"],
            ["24", "Average number of senders per transaction", "Average number of senders per output transaction", "Average"],
            ["25", "Max number of senders in a transaction", "Maximum senders in any single output transaction", "Maximum"],
            ["26", "Min number of senders in a transaction", "Minimum senders in any single output transaction", "Minimum"],
            ["27", "Variation (std. dev.) in senders count", "Standard deviation of senders per output transaction", "Statistics"],
            ["28", "Total receivers addresses in out. transactions", "Total count of all receiver addresses in output transactions", "Count"],
            ["29", "Number of unique receivers addresses", "Count of unique receiver addresses in output transactions", "Count"],
            ["30", "Average number of receivers per transaction", "Average number of receivers per output transaction", "Average"],
            ["31", "Max number of receivers in a transaction", "Maximum receivers in any single output transaction", "Maximum"],
            ["32", "Min number of receivers in a transaction", "Minimum receivers in any single output transaction", "Minimum"],
            ["33", "Variation in receivers count", "Standard deviation of receivers per output transaction", "Statistics"],
            ["34", "Total coins received (excluding change)", "Total BTC amount received from others (change excluded)", "BTC Amount"],
            ["35", "Average coins received per transaction", "Average BTC amount received per output transaction", "BTC Amount"],
            ["36", "Min coins received in one transaction", "Minimum BTC amount received in any single transaction", "BTC Amount"],
            ["37", "Max coins received in one transaction", "Maximum BTC amount received in any single transaction", "BTC Amount"],
            ["38", "Variation (i.e., S.D) in coins received in all transaction", "Standard deviation of BTC amounts received per transaction", "Statistics"],
            ["39", "Avg. coins per sender", "Average BTC amount received per sender address", "BTC Amount"],
        ]
        
        for row_idx, definition in enumerate(definitions, 3):
            for col_idx, value in enumerate(definition, 1):
                cell = definitions_sheet.cell(row=row_idx, column=col_idx, value=value)
                
                if row_idx == 3:  #################### Main header row
                    cell.font = Font(bold=True, color='FFFFFF', size=11)
                    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                #################### Format section headers
                elif "PARAMETERS" in str(value):
                    cell.font = Font(bold=True, color='FFFFFF', size=12)
                    if "INPUT" in str(value):
                        cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
                    else:
                        cell.fill = PatternFill(start_color='E67C73', end_color='E67C73', fill_type='solid')
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    definitions_sheet.row_dimensions[row_idx].height = 25
                #################### Format parameter numbers
                elif col_idx == 1 and str(value).isdigit():
                    cell.font = Font(bold=True, color='1F4E79', size=11)
                    cell.fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')
                #################### Format parameter names (column 2)
                elif col_idx == 2 and str(value) and not "PARAMETERS" in str(value) and value != "":
                    cell.font = Font(bold=True, color='1F4E79', size=10)
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                #################### Format descriptions (column 3)
                elif col_idx == 3 and str(value) and value != "":
                    cell.font = Font(color='333333', size=10)
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                #################### Format categories (column 4)
                elif col_idx == 4 and str(value) and value != "":
                    if "BTC Amount" in str(value):
                        cell.font = Font(bold=True, color='C55A11', size=10)
                    elif "Statistics" in str(value):
                        cell.font = Font(bold=True, color='2E75B6', size=10)
                    elif "Count" in str(value):
                        cell.font = Font(bold=True, color='70AD47', size=10)
                    else:
                        cell.font = Font(bold=True, color='7C7C7C', size=10)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                #################### Regular formatting
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                #################### Add borders to all cells
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
        
        ####################### Summary sheet ###################
        summary_sheet.column_dimensions['A'].width = 35
        summary_sheet.column_dimensions['B'].width = 25
        summary_sheet.column_dimensions['C'].width = 20
        summary_sheet.column_dimensions['D'].width = 20
        summary_sheet.column_dimensions['E'].width = 12
        summary_sheet.column_dimensions['F'].width = 12
        

        detail_sheet.column_dimensions['A'].width = 40  # Bitcoin Address
        # Use get_column_letter to avoid accessing merged cells in row 1 (which
        # yield MergedCell objects without column_letter and can raise errors)
        for col_num in range(2, len(dataframe.columns) + 1):
            col_letter = get_column_letter(col_num)
            detail_sheet.column_dimensions[col_letter].width = 18
        
        definitions_sheet.column_dimensions['A'].width = 8   # Parameter number
        definitions_sheet.column_dimensions['B'].width = 50  # Parameter name
        definitions_sheet.column_dimensions['C'].width = 70  # Description
        definitions_sheet.column_dimensions['D'].width = 15  # Category
        
        summary_sheet.freeze_panes = 'A3'
        detail_sheet.freeze_panes = 'B5'  # Freeze address column and header rows
        definitions_sheet.freeze_panes = 'A4'
        
        #################### Save the workbook ###################
        try:
            wb.save(filename)
            final_name = filename
        except PermissionError:
            from datetime import datetime as _dt
            base, ext = os.path.splitext(filename)
            alt_name = f"{base}_styled_{_dt.now().strftime('%Y%m%d_%H%M%S')}{ext}"
            wb.save(alt_name)
            final_name = alt_name
            print(f" Target file was locked/open. Saved styled report as {final_name}")
        print(f"\n Successfully saved professional Excel report to {final_name}")
        print(f" Report includes 3 sheets:")
        print(f"    Summary - Key statistics and address overview")
        print(f"    Detailed Analysis - All 39 parameters with color coding")
        print(f"    Parameter Guide - Complete definitions and explanations")
        print(f" Features: Professional formatting, color coding, frozen panes, and optimized layouts")
        
    except ImportError:
        print(f"\n️  openpyxl not available. Creating basic Excel file...")
        #################### Fallback to basic Excel ###################
        dataframe.to_excel(filename, index=False)
        print(f" Saved basic Excel file: {filename}")
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            wb_min = load_workbook(filename)
            ws_min = wb_min.active
            for cell in ws_min[1]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            ws_min.freeze_panes = 'B2'
            for col in range(1, ws_min.max_column + 1):
                col_letter = ws_min.cell(row=1, column=col).column_letter
                ws_min.column_dimensions[col_letter].width = 22 if col > 1 else 40
            wb_min.save(filename)
        except Exception:
            pass
    except Exception as e:
        print(f"\n Error creating professional Excel file: {e}")
        try:
            dataframe.to_excel(filename, index=False)
            print(f" Saved basic Excel file as fallback: {filename}")
            try:
                from openpyxl import load_workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                wb_min = load_workbook(filename)
                ws_min = wb_min.active
                for cell in ws_min[1]:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                ws_min.freeze_panes = 'B2'
                for col in range(1, ws_min.max_column + 1):
                    col_letter = ws_min.cell(row=1, column=col).column_letter
                    ws_min.column_dimensions[col_letter].width = 22 if col > 1 else 40
                wb_min.save(filename)
                print(" Applied header styling to the fallback Excel file.")
            except Exception:
                pass
        except Exception as e2:
            print(f" Failed to save even basic Excel: {e2}")

def test_single_address(address):
    """
    Test function to analyze a single address first before running the full analysis.
    This helps verify that the rate limiting fixes are working.
    """
    print(f" Testing with single address: {address}")
    test_df = analyze_addresses([address])
    
    if not test_df.empty:
        print(" Single address test successful!")
        save_to_excel(test_df, f'test_{address[:10]}.xlsx')
        return True
    else:
        print(" Single address test failed.")
        return False

if __name__ == "__main__":
    #################### The main entry point of the script. ###################
    print(" Bitcoin Address Analyzer - Blockchain.info API")
    print("=" * 65)
    
    #################### Validate API connectivity ###################
    print("\n ##############  Testing API connectivity... ###############")
    if validate_api_connectivity():
        print(" Great! ---- Blockchain.info API is accessible ----")
    else:
        print("️  API connectivity test failed - but we'll try anyway")

    print("=" * 65)
    
    #################### Load addresses from file if it exists ###################
    try:
        with open('addresses.txt', 'r') as f:
            file_addresses = [line.strip() for line in f if line.strip()]
        print(f" Found {len(file_addresses)} addresses in addresses.txt")
    except FileNotFoundError:
        file_addresses = []
        print(" No addresses.txt file found")
    
    #################### Get user choice ###################
    choice = get_user_choice()

    if choice == 1:
        #################### Use default addresses ###################
        addresses_to_process = bitcoin_addresses
        print(f"\n Processing {len(addresses_to_process)} default addresses...")
        
    elif choice == 2:
        #################### Use addresses from file ###################
        if file_addresses:
            addresses_to_process = file_addresses
            print(f"\n Processing {len(addresses_to_process)} addresses from addresses.txt")
        else:
            print(" No addresses found in addresses.txt")
            exit(1)
            
    elif choice == 3:
        #################### For single address ###################
        test_address = file_addresses[0] if file_addresses else bitcoin_addresses[0]
        addresses_to_process = [test_address]
        print(f"\n Testing with single address: {test_address}")
        
    else:
        # Exit
        print(" Goodbye!")
        exit(0)
    
    #################### Run the analysis ###################
    try:
        print(f"\n Starting analysis using Blockchain.info API...")
        analysis_df = analyze_addresses(addresses_to_process)
        
        if not analysis_df.empty:
            filename = f'bitcoin_analysis_{len(addresses_to_process)}_addresses.xlsx'
            save_to_excel(analysis_df, filename)
            print(f"\n Analysis completed successfully!")
            print(f" Results saved to {filename}")
            
            #################### Display summary ###################
            print(f"\n Summary:")
            total_balance = analysis_df['Current Balance (BTC)'].sum()
            total_received = analysis_df['Total Received (BTC)'].sum()
            total_sent = analysis_df['Total Sent (BTC)'].sum()
            total_in_txs = analysis_df['1. No. of in. transactions'].sum()
            total_out_txs = analysis_df['21. Number of out. transactions'].sum()
            
            print(f"   Total Balance: {total_balance:.8f} BTC")
            print(f"   Total Received: {total_received:.8f} BTC") 
            print(f"   Total Sent: {total_sent:.8f} BTC")
            print(f"   Total Input Transactions: {total_in_txs}")
            print(f"   Total Output Transactions: {total_out_txs}")
            print(f"   Total Transactions Analyzed: {total_in_txs + total_out_txs}")
        else:
            print(" No data to save. All addresses failed to process.")
            
    except KeyboardInterrupt:
        print("\n\n  Analysis interrupted by user.")
        print(" You can restart the script to try again.")
    except Exception as e:
        print(f"\n Unexpected error: {e}")
        print(" Please check your network connection.")
